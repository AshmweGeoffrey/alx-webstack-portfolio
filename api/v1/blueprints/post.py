from flask import Blueprint, jsonify, request, redirect, session,send_file,url_for,flash
import pandas as pd
from models.inventory_model import inventory
from models.outgoing_stock import outgoing_stock
from models.sale_weekly import sale_weekly
from models.category import category
from models.branch import branch
from models.remark import remark
from . import api
from flask_cors import CORS
from models import storage
from requests import get
import io
import os
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy import create_engine
from models.users import User
from models import mongo_storage
from bcrypt import hashpw, gensalt
import subprocess
import shlex
from pymongo import MongoClient
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
CORS(api)
url='http://localhost:5001'
email_url='http://localhost:5001/activate'
sender_password = "nfxu suvy cdlr hqbe"
sender_email = "axsystem596@gmail.com"
def send_email(sender_email, sender_password, receiver_email, subject, html_body):
    # Create a multipart message
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    message['Subject'] = subject

    # Attach the HTML body
    message.attach(MIMEText(html_body, 'html'))

    # Connect to the server and send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
@api.route('/post/new_item', methods=['POST'])
def post_new_item():
    # Get form data from the request
    name = request.form.get('item-name')
    category = request.form.get('category')
    quantity = request.form.get('quantity')

    # Process the data as needed
    # For example, you can save it to a database

    # Return a JSON response
    response = {
            'name': name,
            'category_name': category,
            'inventory_quantity': int(quantity)
        }
    print(response)
    inv=inventory(**response)
    inv.save()
    flash("New Item Added",category='success')
    return redirect('{}/home'.format(url))
@api.route('/post/outgoing', methods=['POST'])
def post_outgoing():
    # Get form data from the request
    item_name = request.form.get('item-name')
    #category_name = request.form.get('category')
    quantity = request.form.get('new-item-quantity')
    #user_name = request.form.get('user')
    branch_name = request.form.get('Branch')
    user_name = request.form.get('username')
    # Process the data as needed
    # For example, you can save it to a database

    # Return a JSON response
    response = {
            'item_name': item_name,
            #'category_name': category_name,
            'quantity': int(quantity),
            'user_name': user_name,
            'branch_name': branch_name
        }
    print(response)
    out= outgoing_stock(**response)
    out.save()
    flash("Out-Going Item recorded",category='success')
    return redirect('{}/home'.format(url))
@api.route('/post/sale', methods=['POST'])
def post_sale():
    # Get form data from the request
    item_name = request.form.get('item-name')
    quantity = request.form.get('sale-quantity')
    price = request.form.get('sale-price')
    user_name = request.form.get('username')
    payment_method = request.form.get('Payment')
    #branch_name = request.form.get('Branch')

    # Process the data as needed
    # For example, you can save it to a database

    # Return a JSON response
    response = {
            'item_name': item_name,
            'quantity': int(quantity),
            'price': int(price),
            'user_name': user_name,
            'payment_method': payment_method
        }
    sale= sale_weekly(**response)
    sale.save()
    flash("Sale Recorded",category='success')
    return redirect('{}/home'.format(url))
@api.route('/post/existing', methods=['POST'])
def existing_item():
    # Get form data from the request
    item_name = request.form.get('item-name')
    quantity = request.form.get('Existing-quantity')
    #user_name = request.form.get('user')
    # Process the data as needed
    # For example, you can save it to a database
    # Return a JSON response
    response = {
            'item_name': item_name,
            'quantity': int(quantity)
        }
    item_name = item_name.replace("'", "\\'")
    res=storage.command("UPDATE inventory SET inventory_quantity = inventory_quantity+{:d} WHERE name='{}'".format(int(quantity),item_name))
    flash("Item Added Successfully",category='success')
    return redirect('{}/home'.format(url))
@api.route('/post/remark', methods=['POST'])
def remark1():
    message = request.form.get('input5-1')
    response = {
            'message': message
        }
    print(response)
    rem=remark(**response)
    rem.save()
    flash("Report Added",category='success')
    return redirect('{}/home'.format(url))
@api.route('/post/category', methods=['POST'])
def add_category():
    category_1 = request.form.get('category')
    response = {
            'name': category_1,
            'percentage':'0%'
        }
    print(response)
    cat=category(**response)
    cat.save()
    flash("Category Added",category='success')
    return redirect('{}/set_up'.format(url))
@api.route('/post/branch', methods=['POST'])
def add_branch():
    branch_name = request.form.get('branch')
    response = {
            'branch_name': branch_name
        }
    print(response)
    br=branch(**response)
    br.save()
    flash("Branch Added",category='success')
    return redirect('{}/set_up'.format(url))
@api.route('/post/register', methods=['POST'])
def register():
    # Get form data from the request
    user_name = request.form.get('register-username')
    email = request.form.get('register-email')
    password = request.form.get('register-password')
    confrim_password = request.form.get('register-confirm_password')
    access_control = request.form.get('register-Role')
    company = request.form.get('register-Company')
    contact=request.form.get('register-Phone')
    Address=request.form.get('register-Address')
    db_name=("{}_db".format(company))
    # check password are same
    if password != confrim_password:
        flash("Passwords do not match",category='error')
        return redirect('{}/register'.format(url))
    else:
        password = hashpw(password.encode('utf-8'), gensalt())
    # Process the data as needed
    data_new_user = {
        'user_name': user_name,
        'email': email,
        'password': password,
        'access_control': access_control,
        'company': company,
        'contact': contact,
        'address': Address,
        'db_name': db_name,
        'status':'inactive'
    }
    user = User(**data_new_user)
    # create database
    client=MongoClient('localhost', 27017)
    db=client['Ax']
    collection=db['user']
    if collection.find_one({'user_name':user_name}):
        flash("User already exists",category='error')
        return redirect('{}/register'.format(url))
    if collection.find_one({'company':company}): 
        mongo_storage.save(user)
        flash("Company already exists Added User",category='success')
        return redirect('{}/register'.format(url))
    create_database = "CREATE DATABASE IF NOT EXISTS {}".format(db_name)
    mysql_command = f"mysql -uroot -pAshimwe#001 -e {shlex.quote(create_database)}"
    subprocess.run(mysql_command, shell=True, check=True, capture_output=True, text=True)
    print(os.system('mysql -uroot -pAshimwe#001 {} < api/v1/blueprints/database/initial-database'.format(db_name))) 
    mongo_storage.save(user)
    #activating the user
    subject = "Account Activation"
    user_name = data_new_user['user_name']
    # HTML body of the email
    body = """
    <html>
    <head>
    <style>
        /* Fallback fonts since @import may not work in most email clients */
        header {{
            font-family: 'Arial', sans-serif;
            font-size: 24px;
            color: #00BF63;
            border-bottom: 1px solid #00BF63;
            padding-bottom: 10px;
        }}

        h1 {{
            font-family: 'Verdana', sans-serif;
            color: #4CAF50;
        }}

        p {{
            font-family: 'Tahoma', sans-serif;
            font-size: 16px;
        }}
    </style>
    </head>
    <body>
    <header>Ax</header>
        <h2>Hello {}</h2>
        <h3>Activate Your Account</h3>
        <p>Please click the link below to activate your account:</p>
        <p><a href="{}/{}">Activate Account</a></p>
        <p>Thank you!</p>
    </body>
    </html>
    """.format(user_name,email_url,user_name)
    receiver_email = data_new_user['email']
    send_email(sender_email, sender_password, receiver_email, subject, body)
    flash("Account created successfully",category='success')
    return redirect('{}/login'.format(url))
@api.route('/post/login', methods=['GET'])
def login_send():
    res = storage.command("SELECT * FROM user").fetchall()
    users = {user[1]: user[2] for user in res}
    return jsonify(users)
@api.route('/post/download', methods=['POST', 'GET'])
def download():
    excel_file, filename = generate_weekly_report()
    if excel_file and filename:
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    else:
        return jsonify({"error": "Failed to generate report"}), 500
def generate_weekly_report():
    try:
        # Create a database connection
        engine = storage.engine

        # Read the tables into pandas DataFrames
        df_sale_weekly = pd.read_sql_table('sale_weekly', engine)
        df_outgoing_stock = pd.read_sql_table('outgoing_stock', engine)
        df_inventory = pd.read_sql_table('inventory', engine)

        # Create a BytesIO object to store the Excel file
        output = io.BytesIO()

        # Generate a timestamp for the filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "weekly_report_{t}.xlsx".format(t=timestamp)

        # Use pandas to write the DataFrames to an Excel file
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write each DataFrame to a separate sheet
            df_sale_weekly.to_excel(writer, sheet_name='Sale Weekly', index=False)
            df_outgoing_stock.to_excel(writer, sheet_name='Outgoing Stock', index=False)
            df_inventory.to_excel(writer, sheet_name='Inventory', index=False)

            # Get the workbook
            workbook = writer.book

            # Format function
            def format_sheet(worksheet):
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

                # Format header
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                         top=Side(style='thin'), bottom=Side(style='thin'))

            # Apply formatting to each sheet
            format_sheet(workbook['Sale Weekly'])
            format_sheet(workbook['Outgoing Stock'])
            format_sheet(workbook['Inventory'])

        # Save the file locallyi
        save_path = os.path.join('Reports', filename)
        with open(save_path, 'wb') as f:
            f.write(output.getvalue())

        # Seek to the beginning of the stream
        output.seek(0)

        return output, filename

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None, None
